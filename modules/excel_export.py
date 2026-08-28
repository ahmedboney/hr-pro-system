"""تصدير تقارير Excel منسقة باحترافية (xlsx)"""
import os
from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties


# ---------- نظام الألوان الفاخر ----------
COLOR_HEADER = "1E3A8A"        # كحلي عميق
COLOR_HEADER_TEXT = "FFFFFF"
COLOR_TITLE = "1E40AF"
COLOR_SUBTITLE = "475569"
COLOR_STRIPE = "EFF6FF"        # أزرق فاتح للصفوف المخططة
COLOR_TOTAL = "DBEAFE"
COLOR_BORDER = "94A3B8"
COLOR_COMPANY = "1E3A8A"
COLOR_MONEY = "166534"         # أخضر غامق للأموال

THIN_SIDE = Side(style='thin', color=COLOR_BORDER)
MEDIUM_SIDE = Side(style='medium', color=COLOR_BORDER)
BORDER_ALL = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
BORDER_MEDIUM = Border(left=MEDIUM_SIDE, right=MEDIUM_SIDE, top=MEDIUM_SIDE, bottom=MEDIUM_SIDE)

FONT_AR = "Cairo"
FONT_EN = "Calibri"

EMPLOYMENT_TYPES = {
    "full_time": "دوام كامل",
    "part_time": "دوام جزئي",
    "contract": "عقد مؤقت",
}
EMPLOYEE_STATUS = {
    "active": "نشط",
    "on_leave": "في إجازة",
    "terminated": "منتهي",
}


class ExcelExporter:
    """مولّد ملفات Excel باحترافية"""

    def __init__(self, company_name="شركتي المتميزة", subtitle="نظام الموارد البشرية المتكامل",
                 currency="ج.م", by_line="محاسب / أحمد عبدالله"):
        from models import Setting
        self.company = Setting.get('company_name', company_name)
        self.subtitle = subtitle
        self.currency = Setting.get('currency', currency)
        sign_role = Setting.get('sign_fin_role', 'محاسب')
        sign_name = Setting.get('sign_fin_name', 'أحمد عبدالله')
        self.by_line = f"{sign_role} / {sign_name}" if sign_name else sign_role
        self.footer_sign = f"إعداد الأستاذ / {self.by_line}"

    def new_sheet(self, wb, title, right_to_left=True):
        """ننشئ ورقة جديدة مخصصة"""
        ws = wb.create_sheet(title)
        ws.sheet_view.rightToLeft = right_to_left
        ws.sheet_properties = WorksheetProperties(
            pageSetUpPr=PageSetupProperties(fitToPage=True)
        )
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        return ws

    def build_report(self, wb, sheet_name, headers, rows, title_text,
                     period_text=None, total_rows=None, money_columns=(),
                     id_column=None):
        """
        يبني تقريراً كاملاً داخل ورقة جديدة في wb ويعيد الورقة.

        headers: قائمة رؤوس الأعمدة
        rows: قائمة من قوائم القيم (يجب تطابق عددها مع headers)
        total_rows: قوائم اضافية بقيم إجمالية (تظهر كصف مجاميع)
        money_columns: المجموعة التي تشير إلى أعمدة أرقام (تنسيق الالاف)
        """
        ws = self.new_sheet(wb, sheet_name)

        ncols = len(headers)
        last_col = get_column_letter(ncols)

        # 1) عنوان الشركة
        ws.merge_cells(f"A1:{last_col}1")
        c = ws.cell(row=1, column=1, value=self.company)
        c.font = Font(name=FONT_AR, size=18, bold=True, color=COLOR_COMPANY)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 36

        # 2) السطر الفرعي
        ws.merge_cells(f"A2:{last_col}2")
        c = ws.cell(row=2, column=1, value=self.subtitle)
        c.font = Font(name=FONT_AR, size=12, color=COLOR_SUBTITLE)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 22

        # 3) عنوان التقرير
        ws.merge_cells(f"A3:{last_col}3")
        c = ws.cell(row=3, column=1, value=title_text)
        c.font = Font(name=FONT_AR, size=14, bold=True, color=COLOR_TITLE)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = PatternFill('solid', fgColor="F1F5F9")
        ws.row_dimensions[3].height = 28

        # 4) سطر الفترة والتاريخ
        info = f"فترة التقرير: {period_text or '—'}    |    تاريخ الإصدار: {date.today().strftime('%Y-%m-%d')}"
        ws.merge_cells(f"A4:{last_col}4")
        c = ws.cell(row=4, column=1, value=info)
        c.font = Font(name=FONT_AR, size=10, color=COLOR_SUBTITLE)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[4].height = 20

        # 5) رأس الجدول
        header_row = 6
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=h)
            cell.font = Font(name=FONT_AR, size=11, bold=True, color=COLOR_HEADER_TEXT)
            cell.fill = PatternFill('solid', fgColor=COLOR_HEADER)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = BORDER_MEDIUM
        ws.row_dimensions[header_row].height = 26

        # 6) صفوف البيانات
        r = header_row + 1
        for i, row in enumerate(rows):
            fill = PatternFill('solid', fgColor=COLOR_STRIPE) if i % 2 == 0 else None
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r, column=col_idx, value=value)
                cell.border = BORDER_ALL
                cell.font = Font(name=FONT_AR, size=10)
                is_money = col_idx in money_columns
                if is_money and isinstance(value, (int, float)):
                    cell.number_format = f'#,##0.00" {self.currency}"'
                    cell.font = Font(name=FONT_EN, size=10, bold=True, color=COLOR_MONEY)
                    cell.alignment = Alignment(horizontal='center')
                elif id_column and col_idx == id_column:
                    cell.font = Font(name=FONT_EN, size=10, color=COLOR_SUBTITLE)
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
                if fill:
                    cell.fill = fill
            r += 1

        # 7) صفوف المجاميع
        for total_row in (total_rows or []):
            for col_idx, value in enumerate(total_row, start=1):
                cell = ws.cell(row=r, column=col_idx, value=value)
                cell.border = BORDER_MEDIUM
                cell.fill = PatternFill('solid', fgColor=COLOR_TOTAL)
                cell.font = Font(name=FONT_AR, size=11, bold=True)
                if col_idx in money_columns and isinstance(value, (int, float)):
                    cell.number_format = f'#,##0.00" {self.currency}"'
                    cell.font = Font(name=FONT_EN, size=11, bold=True, color=COLOR_MONEY)
                cell.alignment = Alignment(horizontal='center' if col_idx in money_columns else 'right')
            r += 1

        # 8) الفلاتر
        data_end = r - 1
        if rows:
            ws.auto_filter.ref = f"A{header_row}:{last_col}{data_end}"

        # 9) عروض الأعمدة الذكية
        for col_idx in range(1, ncols + 1):
            letter = get_column_letter(col_idx)
            max_len = 8
            for rr in range(header_row, min(data_end, header_row + 60)):
                v = ws.cell(row=rr, column=col_idx).value
                if v is not None:
                    length = len(str(v))
                    if isinstance(v, (int, float)):
                        length = 12
                    max_len = max(max_len, min(length, 45))
            ws.column_dimensions[letter].width = max_len + 4

        ws.freeze_panes = f"A{header_row + 1}"
        ws.print_title_rows = f"{header_row}:{header_row}"
        return ws

    # ---------- ختم وتوقيع التقرير (حقوق الملكية) ----------
    def footer(self, ws, ncols, generated_by="نظام الموارد البشرية المتكامل"):
        last_col = get_column_letter(ncols)
        # توقيع المحاسب
        r = ws.max_row + 2
        ws.merge_cells(f"A{r}:{last_col}{r}")
        c = ws.cell(row=r, column=1, value=self.footer_sign)
        c.font = Font(name=FONT_AR, size=10, bold=True, color=COLOR_SUBTITLE)
        c.alignment = Alignment(horizontal='center')

        r += 1
        ws.merge_cells(f"A{r}:{last_col}{r}")
        c = ws.cell(
            row=r, column=1,
            value=f"إعداد البرنامج: {generated_by} - حقوق الملكية محفوظة © {date.today().year} - {self.by_line}"
        )
        c.font = Font(name=FONT_AR, size=9, bold=True, color=COLOR_COMPANY)
        c.alignment = Alignment(horizontal='center')
        ws.row_dimensions[r].height = 24

        # منطقة التوقيعات
        r += 2
        mid = (ncols + 1) // 2
        left_label = ws.cell(row=r, column=1, value="مدير الموارد البشرية")
        left_label.font = Font(name=FONT_AR, size=10, bold=True)
        left_label.alignment = Alignment(horizontal='center')
        ws.merge_cells(f"A{r}:{get_column_letter(mid)}{r}")

        right_label = ws.cell(row=r, column=mid + 1, value=f"المدير المالي / {self.by_line}")
        right_label.font = Font(name=FONT_AR, size=10, bold=True)
        right_label.alignment = Alignment(horizontal='center')
        ws.merge_cells(f"{get_column_letter(mid + 1)}{r}:{last_col}{r}")

        r += 1
        ws.merge_cells(f"A{r}:{get_column_letter(mid)}{r}")
        c1 = ws.cell(row=r, column=1, value="")
        c1.border = Border(bottom=Side(style='thin'))
        ws.merge_cells(f"{get_column_letter(mid + 1)}{r}:{last_col}{r}")
        c2 = ws.cell(row=r, column=mid + 1, value="")
        c2.border = Border(bottom=Side(style='thin'))
        ws.row_dimensions[r].height = 34

    def finalize(self):
        wb = Workbook()
        wb.remove(wb.active)
        return wb

    def save_bytes(self, wb):
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()


# ==================== تقرير الموظفين الشامل ====================

def export_employees(employees):
    exporter = ExcelExporter(subtitle="نظام الموارد البشرية المتكامل - تقرير الموظفين")
    wb = exporter.finalize()

    headers = ["رقم الموظف", "رقم البصمة", "الاسم", "القسم", "الوظيفة",
               "الهاتف", "الرقم القومي", "تاريخ التعيين", "نوع التعيين",
               "الراتب الأساسي", "البدلات", "إجمالي الراتب", "الحالة"]
    rows = []
    for e in employees:
        rows.append([
            e.emp_id, e.fingerprint_id or "--",
            e.full_name,
            e.department.name if e.department else "—",
            e.position.title if e.position else "—",
            e.phone or "—", e.national_id or "—",
            e.hire_date.strftime("%Y-%m-%d") if e.hire_date else "—",
            EMPLOYMENT_TYPES.get(e.employment_type, e.employment_type),
            float(e.base_salary),
            round(float(e.total_salary) - float(e.base_salary), 2),
            float(e.total_salary),
            EMPLOYEE_STATUS.get(e.status, e.status),
        ])

    money = {10, 11, 12}
    total_basic = sum(float(e.base_salary) for e in employees)
    total_salary = sum(float(e.total_salary) for e in employees)
    total_row = ["", "", "الإجمالي", "", "", "", "", "", "",
                 total_basic, round(total_salary - total_basic, 2), total_salary,
                 f"{len(employees)} موظف"]

    ws = exporter.build_report(
        wb, "الموظفون", headers, rows, "تقرير الموظفين الشامل",
        period_text=f"إجمالي الموظفين: {len(employees)}",
        total_rows=[total_row], money_columns=money, id_column=1,
    )
    exporter.footer(ws, len(headers))
    return exporter.save_bytes(wb)


# ==================== تقرير الحضور الشهري ====================

def export_attendance(month, year, summary):
    exporter = ExcelExporter(subtitle="نظام الموارد البشرية المتكامل - تقرير الحضور والانصراف")
    wb = exporter.finalize()

    months_ar = [
        "يناير", "فبراير", "مارس", "أبريل", "مايو", "يونيو",
        "يوليو", "أغسطس", "سبتمبر", "أكتوبر", "نوفمبر", "ديسمبر",
    ]
    month_name = months_ar[month - 1] if 1 <= month <= 12 else str(month)

    headers = ["رقم الموظف", "الاسم", "القسم", "أيام الحضور", "أيام التأخير",
               "أيام الغياب", "ساعات إضافية", "نسبة الحضور%"]

    rows = []
    for item in summary:
        emp = item['employee']
        rows.append([
            emp.emp_id, emp.full_name,
            emp.department.name if emp.department else "—",
            item['present_days'], item['late_days'], item['absent_days'],
            item['overtime_hours'], item['attendance_rate'],
        ])

    total_present = sum(r[3] for r in rows)
    total_late = sum(r[4] for r in rows)
    total_absent = sum(r[5] for r in rows)
    total_overtime = sum(r[6] for r in rows)
    total_row = ["", "الإجمالي", "", total_present, total_late, total_absent,
                 round(total_overtime, 2), ""]

    ws = exporter.build_report(
        wb, "الحضور الشهري", headers, rows,
        f"تقرير الحضور والانصراف الشهري - {month_name} {year}",
        period_text=f"شهر {month_name} {year}",
        total_rows=[total_row], id_column=1,
    )
    exporter.footer(ws, len(headers))
    return exporter.save_bytes(wb)


# ==================== كشف رواتب الفترة ====================

def export_payroll(period, records):
    exporter = ExcelExporter(subtitle="نظام الموارد البشرية المتكامل - كشف رواتب")
    wb = exporter.finalize()

    headers = ["رقم الموظف", "الاسم", "الراتب الأساسي", "البدلات",
               "إضافي ومكافآت", "الإجمالي", "خصومات", "أقساط السلف",
               "صافي الراتب", "الدفع"]
    rows = []
    for rec in records:
        emp = rec.employee
        bonuses = round(rec.overtime_amount + rec.bonus_amount, 2)
        allowances = round(
            rec.housing_allowance + rec.transport_allowance + rec.food_allowance +
            rec.phone_allowance + rec.other_allowances, 2
        )
        if rec.status in ('paid', 'closed'):
            pay_status = "مدفوع"
        else:
            pay_status = "غير مدفوع"
        rows.append([
            emp.emp_id, emp.full_name,
            round(rec.base_salary, 2), allowances, bonuses,
            round(rec.gross_salary, 2), round(rec.total_deductions - rec.loan_deduction, 2),
            round(rec.loan_deduction, 2), round(rec.net_salary, 2), pay_status,
        ])

    money = {3, 4, 5, 6, 7, 8, 9}
    t_basic = sum(r[2] for r in rows)
    t_allow = sum(r[3] for r in rows)
    t_bonus = sum(r[4] for r in rows)
    t_gross = sum(r[5] for r in rows)
    t_ded = sum(r[6] for r in rows)
    t_loan = sum(r[7] for r in rows)
    t_net = sum(r[8] for r in rows)
    total_row = ["", "الإجمالي", t_basic, t_allow, t_bonus, t_gross, t_ded, t_loan, t_net,
                 f"{len(records)} موظف"]

    ws = exporter.build_report(
        wb, "كشف الرواتب", headers, rows,
        f"كشف رواتب فترة {period.name}",
        period_text=period.name,
        total_rows=[total_row], money_columns=money, id_column=1,
    )
    exporter.footer(ws, len(headers))
    return exporter.save_bytes(wb)


# ==================== تقرير أرصدة الإجازات ====================

def export_leave_balances(year, balances_data, leave_types):
    exporter = ExcelExporter(subtitle="نظام الموارد البشرية المتكامل - أرصدة الإجازات")
    wb = exporter.finalize()

    headers = ["رقم الموظف", "الاسم"] + [lt.name for lt in leave_types] + ["إجمالي المتبقي"]
    rows = []
    for item in balances_data:
        emp = item['employee']
        by_type = {b.leave_type_id: b for b in item['balances']}
        row = [emp.emp_id, emp.full_name]
        total_remaining = 0
        for lt in leave_types:
            b = by_type.get(lt.id)
            if b:
                row.append(f"{b.remaining_days} / {b.entitled_days}")
                total_remaining += b.remaining_days
            else:
                row.append("—")
        row.append(total_remaining)
        rows.append(row)

    ws = exporter.build_report(
        wb, "أرصدة الإجازات", headers, rows,
        f"تقرير أرصدة الإجازات - سنة {year}",
        period_text=f"سنة {year}",
        id_column=1,
    )
    exporter.footer(ws, len(headers))
    return exporter.save_bytes(wb)


# ==================== تقرير السلف والقروض ====================

def export_loans(loans):
    exporter = ExcelExporter(subtitle="نظام الموارد البشرية المتكامل - السلف والقروض")
    wb = exporter.finalize()

    headers = ["رقم الموظف", "الاسم", "النوع", "المبلغ", "الأقساط",
               "القسط الشهري", "المتبقي", "الحالة", "السبب"]
    rows = []
    for loan in loans:
        emp = loan.employee
        loan_type = "سلفة" if loan.loan_type == 'savings' else "قرض"
        status_map = {"active": "نشطة", "paid": "مسددة", "cancelled": "ملغاة"}
        rows.append([
            emp.emp_id, emp.full_name, loan_type,
            round(loan.amount, 2), loan.installment_count,
            round(loan.installment_amount or 0, 2),
            round(loan.remaining_amount or 0, 2),
            status_map.get(loan.status, loan.status),
            loan.reason or "—",
        ])

    money = {4, 6, 7}
    t_amount = sum(r[3] for r in rows)
    t_remaining = sum(r[6] for r in rows)
    total_row = ["", "الإجمالي", "", t_amount, "", "", t_remaining, "", ""]

    ws = exporter.build_report(
        wb, "السلف والقروض", headers, rows, "تقرير السلف والقروض",
        period_text=f"إجمالي السجلات: {len(loans)}",
        total_rows=[total_row], money_columns=money, id_column=1,
    )
    exporter.footer(ws, len(headers))
    return exporter.save_bytes(wb)


# ==================== كشف الراتب الفردي ====================

def export_payslip(record):
    """كشف راتب فردي منسق بفخامة (يختلف تخطيطه عن الجداول)"""
    exporter = ExcelExporter(subtitle="كشف راتب موظف")
    emp = record.employee
    wb = exporter.finalize()
    ws = wb.create_sheet("كشف الراتب")
    ws.sheet_view.rightToLeft = True
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1

    ncols = 4
    last_col = get_column_letter(ncols)

    # العناوين
    ws.merge_cells(f"A1:{last_col}1")
    c = ws.cell(row=1, column=1, value=exporter.company)
    c.font = Font(name=FONT_AR, size=18, bold=True, color=COLOR_COMPANY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    ws.merge_cells(f"A2:{last_col}2")
    c = ws.cell(row=2, column=1, value=f"كشف راتب شهر {record.period.name}")
    c.font = Font(name=FONT_AR, size=14, bold=True, color=COLOR_TITLE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill('solid', fgColor="F1F5F9")
    ws.row_dimensions[2].height = 28

    def section(row, label):
        ws.merge_cells(f"A{row}:{last_col}{row}")
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(name=FONT_AR, size=11, bold=True, color=COLOR_HEADER_TEXT)
        c.fill = PatternFill('solid', fgColor=COLOR_HEADER)
        c.alignment = Alignment(horizontal='center')

    def line(row, label, value, money=False, bold=False):
        a = ws.cell(row=row, column=1, value=label)
        b = ws.cell(row=row, column=2, value=value)
        a.font = Font(name=FONT_AR, size=11, bold=True)
        a.alignment = Alignment(horizontal='right')
        b.alignment = Alignment(horizontal='center')
        b.border = BORDER_ALL
        a.border = BORDER_ALL
        if money:
            b.number_format = f'#,##0.00" {exporter.currency}"'
            b.font = Font(name=FONT_EN, size=11, bold=bold, color=COLOR_MONEY)
        else:
            b.font = Font(name=FONT_AR, size=11, bold=bold)

    r = 4
    c = ws.cell(row=r, column=1, value=f"الموظف: {emp.full_name}")
    c.font = Font(name=FONT_AR, size=12, bold=True)
    c = ws.cell(row=r, column=3, value=f"الرقم الوظيفي: {emp.emp_id}")
    c.font = Font(name=FONT_AR, size=11)
    r += 1
    c = ws.cell(row=r, column=1, value=f"القسم: {emp.department.name if emp.department else '—'}")
    c.font = Font(name=FONT_AR, size=11)
    c = ws.cell(row=r, column=3, value=f"الوظيفة: {emp.position.title if emp.position else '—'}")
    c.font = Font(name=FONT_AR, size=11)
    r += 2

    section(r, "بنود الاستحقاق")
    r += 1
    lines = [
        ("الراتب الأساسي", record.base_salary, True),
        ("بدل سكن", record.housing_allowance, True),
        ("بدل انتقال", record.transport_allowance, True),
        ("بدل وجبات", record.food_allowance, True),
        ("بدل هاتف", record.phone_allowance, True),
        ("بدلات أخرى", record.other_allowances, True),
        ("الساعات الإضافية", record.overtime_amount, True),
        ("المكافآت", record.bonus_amount, True),
    ]
    for label, val, m in lines:
        line(r, label, val, money=m)
        r += 1
    line(r, "إجمالي الاستحقاقات", record.gross_salary, money=True, bold=True)
    ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor=COLOR_TOTAL)
    r += 2

    section(r, "الخصومات")
    r += 1
    lines = [
        ("خصم غياب", record.absent_deduction),
        ("خصم تأخير", record.late_deduction),
        ("إجازة بدون أجر", record.unpaid_leave_deduction),
        ("تأمينات اجتماعية", record.social_insurance),
        ("ضريبة كسب العمل", record.tax_amount),
        ("أقساط السلف", record.loan_deduction),
        ("خصومات أخرى", record.other_deductions),
    ]
    for label, val in lines:
        line(r, label, val, money=True)
        r += 1
    line(r, "إجمالي الخصومات", record.total_deductions, money=True, bold=True)
    ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor=COLOR_TOTAL)
    r += 2

    # الصافي
    ws.merge_cells(f"A{r}:{last_col}{r}")
    c = ws.cell(row=r, column=1, value=f"صافي الراتب المستحق")
    c.font = Font(name=FONT_AR, size=14, bold=True)
    c.alignment = Alignment(horizontal='center')
    c.fill = PatternFill('solid', fgColor="FEF9C3")
    c.border = BORDER_MEDIUM
    r += 1
    ws.merge_cells(f"A{r}:{last_col}{r}")
    c = ws.cell(row=r, column=1, value=f"{record.net_salary:,.2f} {exporter.currency}")
    c.font = Font(name=FONT_EN, size=20, bold=True, color=COLOR_MONEY)
    c.alignment = Alignment(horizontal='center')
    c.fill = PatternFill('solid', fgColor="FEF9C3")
    c.border = BORDER_MEDIUM
    ws.row_dimensions[r].height = 34
    r += 2

    for col_idx in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 26

    exporter.footer(ws, ncols)
    return exporter.save_bytes(wb)


# ==================== كشف التأمينات والضرائب ====================

def export_insurance_tax(period, records):
    exporter = ExcelExporter(subtitle="نظام الموارد البشرية المتكامل - كشف التأمينات والضرائب")
    wb = exporter.finalize()

    headers = ["رقم الموظف", "الاسم", "المرتب التأميني",
               "تأمينات (خصم)", "صافي الاستقطاع", "ضريبة كسب العمل"]
    rows = []
    total_insur_salary = 0
    total_insur = 0
    total_tax = 0
    for rec in records:
        emp = rec.employee
        insur_salary = round(rec.base_salary + rec.housing_allowance + rec.transport_allowance + rec.food_allowance, 2)
        rows.append([
            emp.emp_id, emp.full_name, insur_salary,
            round(rec.social_insurance, 2), round(rec.social_insurance, 2),
            round(rec.tax_amount, 2),
        ])
        total_insur_salary += insur_salary
        total_insur += rec.social_insurance
        total_tax += rec.tax_amount

    money = {3, 4, 5, 6}
    total_row = ["", "الإجمالي", round(total_insur_salary, 2),
                 round(total_insur, 2), round(total_insur, 2), round(total_tax, 2)]

    ws = exporter.build_report(
        wb, "التأمينات والضرائب", headers, rows,
        f"كشف التأمينات الاجتماعية وضريبة كسب العمل - فترة {period.name}",
        period_text=period.name,
        total_rows=[total_row], money_columns=money, id_column=1,
    )
    exporter.footer(ws, len(headers))
    return exporter.save_bytes(wb)